import csv
import os
import logging

import pdfplumber
import openpyxl

from app.indexing.documents import FinancialDocument

logger = logging.getLogger(__name__)

SUPPORTED_EXTS = {".pdf", ".xlsx", ".xls", ".csv"}
CHUNK_ROWS = 100  # rows per spreadsheet / CSV chunk


class LocalFinancialParser:
    """Parser for financial documents stored on the local filesystem.

    Walks a directory tree and converts PDFs, spreadsheets, and CSVs into
    FinancialDocument chunks suitable for indexing and retrieval.

    Args:
        local_path: Absolute path to the directory (or single file) to parse.

    Example:
        parser = LocalFinancialParser("/data/finance/annual_reports")
        docs = parser.parse()
    """

    def __init__(self, local_path: str):
        self.local_path = local_path

    def parse(self) -> list[FinancialDocument]:
        """Walk the directory and parse all supported financial documents."""
        documents: list[FinancialDocument] = []
        for file_path in self._walk():
            ext = os.path.splitext(file_path)[1].lower()
            try:
                if ext == ".pdf":
                    documents.extend(self._parse_pdf(file_path))
                elif ext in (".xlsx", ".xls"):
                    documents.extend(self._parse_spreadsheet(file_path))
                elif ext == ".csv":
                    documents.extend(self._parse_csv(file_path))
            except Exception:
                logger.warning(f"Skipping {file_path}: failed to parse", exc_info=True)
        return documents

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _walk(self) -> list[str]:
        """Return sorted list of supported file paths under local_path."""
        if os.path.isfile(self.local_path):
            return [self.local_path]
        paths = []
        for root, _, files in os.walk(self.local_path):
            for fname in sorted(files):
                if os.path.splitext(fname)[1].lower() in SUPPORTED_EXTS:
                    paths.append(os.path.join(root, fname))
        return paths

    def _parse_pdf(self, file_path: str) -> list[FinancialDocument]:
        """One FinancialDocument per page; skip blank pages."""
        docs: list[FinancialDocument] = []
        with pdfplumber.open(file_path) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                text = (page.extract_text() or "").strip()
                if not text:
                    continue
                docs.append(FinancialDocument(
                    text=text,
                    source=file_path,
                    doc_type="pdf",
                    page_or_sheet=f"page {i}",
                ))
        return docs

    def _parse_spreadsheet(self, file_path: str) -> list[FinancialDocument]:
        """Chunk each sheet into CHUNK_ROWS rows; prepend header on continuation chunks."""
        docs: list[FinancialDocument] = []
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [
                "\t".join("" if cell is None else str(cell) for cell in row)
                for row in ws.iter_rows(values_only=True)
                if any(cell is not None for cell in row)
            ]
            if not rows:
                continue
            header = rows[0]
            for start in range(0, len(rows), CHUNK_ROWS):
                chunk = rows[start:start + CHUNK_ROWS]
                if start > 0:
                    chunk = [header] + chunk
                docs.append(FinancialDocument(
                    text="\n".join(chunk),
                    source=file_path,
                    doc_type="spreadsheet",
                    page_or_sheet=f"{sheet_name} rows {start + 1}-{start + len(chunk)}",
                ))
        wb.close()
        return docs

    def _parse_csv(self, file_path: str) -> list[FinancialDocument]:
        """Chunk CSV into CHUNK_ROWS rows; prepend header on continuation chunks."""
        docs: list[FinancialDocument] = []
        with open(file_path, newline="", encoding="utf-8", errors="replace") as f:
            rows = ["\t".join(row) for row in csv.reader(f)]
        if not rows:
            return docs
        header = rows[0]
        for start in range(0, len(rows), CHUNK_ROWS):
            chunk = rows[start:start + CHUNK_ROWS]
            if start > 0:
                chunk = [header] + chunk
            docs.append(FinancialDocument(
                text="\n".join(chunk),
                source=file_path,
                doc_type="csv",
                page_or_sheet=f"rows {start + 1}-{start + len(chunk)}",
            ))
        return docs
