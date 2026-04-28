"""Generate sample financial documents for AcmeWackoWidgets."""
from pathlib import Path
from fpdf import FPDF, XPos, YPos
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

OUT = Path(__file__).parent / "sample_financial_data"

NL = {"new_x": XPos.LMARGIN, "new_y": YPos.NEXT}


def generate_pdf():
    pdf = FPDF()
    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    # Header
    pdf.set_font("Helvetica", "B", 22)
    pdf.cell(0, 12, "AcmeWackoWidgets, Inc.", **NL, align="C")
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 8, "Annual Financial Report - Fiscal Year 2025", **NL, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, "Confidential - Internal Use Only", **NL, align="C")
    pdf.ln(6)

    pdf.set_draw_color(60, 60, 60)
    pdf.set_line_width(0.5)
    pdf.line(20, pdf.get_y(), 190, pdf.get_y())
    pdf.ln(6)

    # Executive Summary
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 7, "Executive Summary", **NL)
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(0, 5.5,
        "AcmeWackoWidgets delivered strong performance in FY 2025, driven by "
        "exceptional demand across the WackoWidget product line. Total consolidated "
        "revenue reached $8,412,300, representing a 14% increase over FY 2024. "
        "Operating income improved to $1,102,450, and net income attributable to "
        "shareholders was $874,200."
    )
    pdf.ln(4)

    # Product Revenue Table
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 7, "Product Revenue Breakdown", **NL)

    headers = ["Product", "Units Sold", "Avg. Unit Price", "Total Revenue"]
    col_widths = [65, 30, 40, 35]

    pdf.set_fill_color(230, 230, 230)
    pdf.set_font("Helvetica", "B", 10)
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 7, h, border=1, fill=True,
                 align="C" if i > 0 else "L")
    pdf.ln()

    rows = [
        ("WackoWidget1000",   "4,210", "$312.00",   "$1,313,520"),
        ("WackoWidget2000",   "3,875", "$588.50",   "$2,280,438"),
        ("WackoWidget3000",   "2,106", "$586.50",   "$1,234,567"),
        ("WackoWidget Pro",   "1,540", "$892.00",   "$1,373,680"),
        ("WackoWidget Elite", "1,012", "$2,184.00", "$2,210,208"),
    ]

    pdf.set_font("Helvetica", "", 10)
    for r in rows:
        for i, cell in enumerate(r):
            pdf.cell(col_widths[i], 6.5, cell, border=1,
                     align="C" if i > 0 else "L")
        pdf.ln()

    # Total row
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(200, 200, 200)
    for i, cell in enumerate(["Total", "12,743", "-", "$8,412,413"]):
        pdf.cell(col_widths[i], 7, cell, border=1, fill=True,
                 align="C" if i > 0 else "L")
    pdf.ln(6)

    # Notable callout
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(255, 255, 200)
    pdf.cell(0, 7,
        "  KEY HIGHLIGHT: WackoWidget3000 generated $1,234,567 in sales for FY 2025,",
        border=1, fill=True, **NL)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6,
        "  surpassing its $900,000 target and achieving 137% of plan.",
        border=1, fill=True, **NL)
    pdf.ln(5)

    # Operating expenses
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 7, "Operating Expenses", **NL)
    exp_rows = [
        ("Cost of Goods Sold",         "$4,206,150"),
        ("Research & Development",     "$1,050,000"),
        ("Sales & Marketing",          "$840,000"),
        ("General & Administrative",   "$1,213,700"),
        ("Total Operating Expenses",   "$7,309,850"),
    ]
    for label, value in exp_rows:
        bold = label.startswith("Total")
        pdf.set_font("Helvetica", "B" if bold else "", 10)
        pdf.cell(130, 6, label)
        pdf.cell(40, 6, value, align="R", **NL)
    pdf.ln(4)

    # Outlook
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 7, "FY 2026 Outlook", **NL)
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(0, 5.5,
        "Management expects continued growth in the WackoWidget3000 and WackoWidget Pro "
        "segments. Revenue guidance for FY 2026 is set at $9.8M-$10.2M. The company "
        "plans to launch the WackoWidget Ultra in Q3 2026, targeting the premium market."
    )
    pdf.ln(4)

    pdf.line(20, pdf.get_y(), 190, pdf.get_y())
    pdf.ln(3)
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 5,
        "AcmeWackoWidgets, Inc.  |  123 Wacky Way, Gadgetville, CA 94000  |  "
        "Copyright 2025 AcmeWackoWidgets. All rights reserved.",
        align="C")

    out_path = OUT / "AcmeWackoWidgets_FY2025_Annual_Report.pdf"
    pdf.output(str(out_path))
    print(f"PDF written: {out_path}")


def generate_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "WackoWidget3000 Sales"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="D9E1F2")
    center = Alignment(horizontal="center")

    # Title
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value = "AcmeWackoWidgets - WackoWidget3000 Quarterly Sales"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:C2")
    ws["A2"].value = "Fiscal Year 2025"
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = center

    # Headers
    ws["A4"].value = "Quarter"
    ws["B4"].value = "Unit Sales (qty)"
    ws["C4"].value = "Revenue ($)"
    for col in ["A", "B", "C"]:
        cell = ws[f"{col}4"]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Data rows - revenues sum to exactly $123,456
    data = [
        ("Q1 FY2025", 142, 27_890),
        ("Q2 FY2025", 158, 31_204),
        ("Q3 FY2025", 163, 33_102),
        ("Q4 FY2025", 152, 31_260),
    ]
    assert sum(r[2] for r in data) == 123_456, "Rows must sum to 123,456"

    for i, (quarter, units, revenue) in enumerate(data, start=5):
        ws[f"A{i}"].value = quarter
        ws[f"B{i}"].value = units
        ws[f"C{i}"].value = revenue
        ws[f"B{i}"].alignment = center
        ws[f"C{i}"].number_format = '$#,##0'

    # Total row
    total_row = len(data) + 5
    ws[f"A{total_row}"].value = "Total"
    ws[f"B{total_row}"].value = f"=SUM(B5:B{total_row - 1})"
    ws[f"C{total_row}"].value = f"=SUM(C5:C{total_row - 1})"
    for col in ["A", "B", "C"]:
        cell = ws[f"{col}{total_row}"]
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center
    ws[f"C{total_row}"].number_format = '$#,##0'

    note_row = total_row + 2
    ws.merge_cells(f"A{note_row}:C{note_row}")
    ws[f"A{note_row}"].value = (
        "Note: Revenue figures reflect net sales after returns and allowances."
    )
    ws[f"A{note_row}"].font = Font(italic=True, size=9)

    out_path = OUT / "WackoWidget3000_Quarterly_Sales.xlsx"
    wb.save(str(out_path))
    print(f"XLSX written: {out_path}")


if __name__ == "__main__":
    generate_pdf()
    generate_xlsx()
